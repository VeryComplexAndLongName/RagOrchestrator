from __future__ import annotations

import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZipFile

from ragflow_orchestrator.templates.utils import extract_text_from_html


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return _read_text_with_fallbacks(path)
    if suffix == ".html":
        return extract_text_from_html(_read_text_with_fallbacks(path))
    if suffix == ".docx":
        return _extract_docx(path)
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".xlsx":
        return _extract_xlsx(path)
    raise ValueError(f"Unsupported extension: {suffix}")


def _extract_docx(path: Path) -> str:
    with ZipFile(path, "r") as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    texts = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
    return "\n".join(texts)


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Install pypdf to extract .pdf files") from exc

    reader = PdfReader(str(path))
    parts = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(parts)


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Install openpyxl to extract .xlsx files") from exc

    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    cells: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None]
            if values:
                cells.append(" | ".join(values))
    return "\n".join(cells)


def _read_text_with_fallbacks(path: Path) -> str:
    data = path.read_bytes()
    encodings = ("utf-8-sig", "utf-8", "cp1251", "cp866", "utf-16")
    for encoding in encodings:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")
