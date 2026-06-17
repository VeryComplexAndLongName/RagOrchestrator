from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from ragflow_orchestrator.chunking.fixed import FixedWindowChunker
from ragflow_orchestrator.chunking.markdown import MarkdownHeadingChunker
from ragflow_orchestrator.models import BaseChunk


class DocumentType(str, Enum):
    CODE = "code"
    PDF = "pdf"
    DOCX = "docx"
    XLSX = "xlsx"
    HTML = "html"
    MARKDOWN = "markdown"
    JSON = "json"
    CSV = "csv"
    XML = "xml"
    TXT = "txt"
    UNSUPPORTED = "unsupported"


@dataclass(slots=True)
class DocumentDetection:
    document_type: DocumentType
    mime_type: str = ""
    source: str = "heuristic"


_EXTENSION_MAP: dict[str, DocumentType] = {
    ".py": DocumentType.CODE,
    ".js": DocumentType.CODE,
    ".ts": DocumentType.CODE,
    ".tsx": DocumentType.CODE,
    ".go": DocumentType.CODE,
    ".java": DocumentType.CODE,
    ".cs": DocumentType.CODE,
    ".rb": DocumentType.CODE,
    ".rs": DocumentType.CODE,
    ".c": DocumentType.CODE,
    ".cpp": DocumentType.CODE,
    ".h": DocumentType.CODE,
    ".md": DocumentType.MARKDOWN,
    ".markdown": DocumentType.MARKDOWN,
    ".html": DocumentType.HTML,
    ".htm": DocumentType.HTML,
    ".docx": DocumentType.DOCX,
    ".xlsx": DocumentType.XLSX,
    ".pdf": DocumentType.PDF,
    ".txt": DocumentType.TXT,
    ".json": DocumentType.JSON,
    ".jsonl": DocumentType.JSON,
    ".csv": DocumentType.CSV,
    ".xml": DocumentType.XML,
}

_CONTENT_TYPE_MAP: dict[str, DocumentType] = {
    "application/pdf": DocumentType.PDF,
    "application/json": DocumentType.JSON,
    "application/xml": DocumentType.XML,
    "text/xml": DocumentType.XML,
    "text/csv": DocumentType.CSV,
    "text/plain": DocumentType.TXT,
    "text/markdown": DocumentType.MARKDOWN,
    "text/html": DocumentType.HTML,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": DocumentType.DOCX,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": DocumentType.XLSX,
    "application/vnd.ms-excel": DocumentType.XLSX,
}

_MARKDOWN_HINTS = ("\n# ", "\n## ", "\n### ", "\n- ", "\n* ", "\n1. ", "```", "|")


class DocumentAwareCleaner:
    """Gentle cleaner that preserves line structure for downstream chunking."""

    _control_chars = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
    _spaces = re.compile(r"[ \t]+")

    def clean(self, text: str) -> str:
        if not text:
            return ""

        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        normalized = self._control_chars.sub("", normalized)
        lines = [self._spaces.sub(" ", line).rstrip() for line in normalized.split("\n")]
        normalized = "\n".join(lines)
        normalized = re.sub(r"\n{3,}", "\n\n", normalized)
        return normalized.strip()


class MarkdownAstChunker:
    def __init__(self, fallback: MarkdownHeadingChunker | None = None) -> None:
        self._fallback = fallback or MarkdownHeadingChunker()

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        parser = self._load_parser()
        if parser is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        try:
            tokens = parser.parse(text)
        except Exception:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        buffer: list[str] = []
        heading_stack: list[str] = []
        pending_heading_level = 0

        def flush() -> None:
            nonlocal index
            if not buffer:
                return
            body = "\n".join(buffer).strip()
            buffer.clear()
            if not body:
                return
            chunk_meta = dict(metadata)
            if heading_stack:
                chunk_meta["heading_path"] = " > ".join(heading_stack)
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=body,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        for token in tokens:
            token_type = getattr(token, "type", "")
            if token_type == "heading_open":
                flush()
                tag = str(getattr(token, "tag", "h1"))
                pending_heading_level = int(tag[1:]) if len(tag) > 1 and tag[1:].isdigit() else 1
                continue

            if token_type == "inline" and pending_heading_level:
                content = str(getattr(token, "content", "")).strip()
                if content:
                    while len(heading_stack) >= pending_heading_level:
                        heading_stack.pop()
                    heading_stack.append(content)
                continue

            if token_type == "heading_close":
                pending_heading_level = 0
                continue

            if token_type in {"fence", "code_block"}:
                flush()
                content = str(getattr(token, "content", "")).strip()
                if content:
                    chunk_meta = dict(metadata)
                    if heading_stack:
                        chunk_meta["heading_path"] = " > ".join(heading_stack)
                    chunk_meta["block_type"] = "code"
                    chunks.append(
                        BaseChunk(
                            id=f"{source_id}:{index}",
                            text=content,
                            metadata=chunk_meta,
                            source_id=source_id,
                            chunk_index=index,
                            created_at=now,
                        )
                    )
                    index += 1
                continue

            if token_type == "inline":
                content = str(getattr(token, "content", "")).strip()
                if content:
                    buffer.append(content)
                continue

            if token_type == "paragraph_close":
                buffer.append("")

        flush()
        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _load_parser() -> Any | None:
        try:
            from markdown_it import MarkdownIt  # type: ignore[import-not-found]
        except ImportError:
            return None

        return MarkdownIt("commonmark", options_update={"html": True})


def _metadata_text(metadata: dict[str, object], *keys: str) -> str:
    for key in keys:
        value = metadata.get(key)
        if value:
            return str(value)
    return ""


def _metadata_path(metadata: dict[str, object], *keys: str) -> Path | None:
    for key in keys:
        value = metadata.get(key)
        if value:
            try:
                return Path(str(value))
            except Exception:
                continue
    return None


def _render_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    if not headers:
        return ""
    width = len(headers)
    normalized_rows = [row[:width] + [""] * max(0, width - len(row)) for row in rows]
    lines = [" | ".join(headers)]
    lines.append(" | ".join("---" for _ in headers))
    for row in normalized_rows:
        lines.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(lines)


class HTMLDOMChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_block_chars: int = 1600) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_block_chars = max_block_chars

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        html_source = _metadata_text(metadata, "_html_source", "html_source", "raw_html") or text
        soup = self._load_bs4(html_source)
        if soup is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        body = getattr(soup, "body", None) or soup
        title = soup.title.get_text(" ", strip=True) if getattr(soup, "title", None) else ""
        heading_stack: list[str] = [title] if title else []
        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        buffer: list[str] = []
        current_block_type = "text"

        def flush() -> None:
            nonlocal index, current_block_type
            if not buffer:
                return
            body_text = "\n".join(buffer).strip()
            buffer.clear()
            if not body_text:
                return
            chunk_meta = dict(metadata)
            if heading_stack:
                chunk_meta["heading_path"] = " > ".join([part for part in heading_stack if part])
            chunk_meta["block_type"] = current_block_type
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=body_text,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1
            current_block_type = "text"

        def append_block(block_text: str, block_type: str) -> None:
            nonlocal current_block_type
            if not block_text.strip():
                return
            if buffer and (len("\n".join(buffer)) + len(block_text) > self._max_block_chars or current_block_type != block_type):
                flush()
            current_block_type = block_type
            buffer.append(block_text.strip())
            if len("\n".join(buffer)) >= self._max_block_chars:
                flush()

        def walk(node: Any) -> None:
            for child in getattr(node, "children", []) or []:
                name = str(getattr(child, "name", "") or "").lower()
                if not name:
                    text_value = str(getattr(child, "string", "") or "").strip()
                    if text_value:
                        append_block(text_value, "text")
                    continue

                if name in {"script", "style", "noscript"}:
                    continue

                if name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
                    flush()
                    level = int(name[1])
                    while len(heading_stack) >= level:
                        heading_stack.pop()
                    heading_text = getattr(child, "get_text", lambda *args, **kwargs: "")(" ", strip=True)
                    if heading_text:
                        heading_stack.append(heading_text)
                    continue

                if name == "table":
                    flush()
                    table_text = self._render_table(child)
                    append_block(table_text, "table")
                    flush()
                    continue

                if name in {"pre", "code"}:
                    flush()
                    code_text = getattr(child, "get_text", lambda *args, **kwargs: "")("\n", strip=True)
                    append_block(code_text, "code")
                    flush()
                    continue

                if name in {"ul", "ol"}:
                    flush()
                    list_text = self._render_list(child)
                    append_block(list_text, "list")
                    flush()
                    continue

                if name in {"p", "blockquote", "li", "td", "th"}:
                    text_value = getattr(child, "get_text", lambda *args, **kwargs: "")(" ", strip=True)
                    if text_value:
                        append_block(text_value, name)
                    continue

                if name in {"section", "article", "main", "div", "body", "header", "footer", "aside"}:
                    walk(child)
                    continue

                nested_text = getattr(child, "get_text", lambda *args, **kwargs: "")(" ", strip=True)
                if nested_text:
                    append_block(nested_text, name)

        walk(body)
        flush()
        return chunks or self._fallback.chunk(source_id=source_id, text=extract_text_from_html(html_source), metadata=metadata)

    @staticmethod
    def _load_bs4(html_source: str) -> Any | None:
        try:
            from bs4 import BeautifulSoup  # type: ignore[import-not-found]
        except ImportError:
            return None

        return BeautifulSoup(html_source, "html.parser")

    @staticmethod
    def _render_list(node: Any) -> str:
        items: list[str] = []
        for child in getattr(node, "find_all", lambda *args, **kwargs: [])("li", recursive=False):
            text_value = getattr(child, "get_text", lambda *args, **kwargs: "")(" ", strip=True)
            if text_value:
                items.append(f"- {text_value}")
        return "\n".join(items)

    @staticmethod
    def _render_table(node: Any) -> str:
        rows: list[list[str]] = []
        for row in getattr(node, "find_all", lambda *args, **kwargs: [])("tr"):
            cells = row.find_all(["th", "td"]) if hasattr(row, "find_all") else []
            values = [cell.get_text(" ", strip=True) for cell in cells if cell.get_text(" ", strip=True)]
            if values:
                rows.append(values)
        if not rows:
            return getattr(node, "get_text", lambda *args, **kwargs: "")(" ", strip=True)
        headers = rows[0]
        return _render_markdown_table(headers, rows[1:])


class DOCXStructureChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_block_chars: int = 2400) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_block_chars = max_block_chars

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        path = _metadata_path(metadata, "file_path")
        if path is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        document = self._load_document(path)
        if document is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        buffer: list[str] = []
        heading_stack: list[str] = []
        current_block_type = "paragraph"

        def flush() -> None:
            nonlocal index, current_block_type
            if not buffer:
                return
            body_text = "\n\n".join(buffer).strip()
            buffer.clear()
            if not body_text:
                return
            chunk_meta = dict(metadata)
            if heading_stack:
                chunk_meta["heading_path"] = " > ".join([part for part in heading_stack if part])
            chunk_meta["block_type"] = current_block_type
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=body_text,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1
            current_block_type = "paragraph"

        def append_block(block_text: str, block_type: str) -> None:
            nonlocal current_block_type
            block_text = block_text.strip()
            if not block_text:
                return
            if buffer and (len("\n\n".join(buffer)) + len(block_text) > self._max_block_chars or current_block_type != block_type):
                flush()
            current_block_type = block_type
            buffer.append(block_text)
            if len("\n\n".join(buffer)) >= self._max_block_chars:
                flush()

        for block in _iter_docx_blocks(document):
            if getattr(block, "is_paragraph", False):
                paragraph = block.paragraph
                text_value = paragraph.text.strip()
                if not text_value:
                    continue
                style_name = str(getattr(getattr(paragraph, "style", None), "name", "") or "")
                heading_level = _heading_level_from_style(style_name)
                if heading_level is not None:
                    flush()
                    while len(heading_stack) >= heading_level:
                        heading_stack.pop()
                    heading_stack.append(text_value)
                    continue
                append_block(text_value, "paragraph")
                continue

            table_text = self._render_table(block.table)
            if table_text.strip():
                flush()
                append_block(table_text, "table")
                flush()

        flush()
        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _load_document(path: Path) -> Any | None:
        try:
            from docx import Document  # type: ignore[import-not-found]
        except ImportError:
            return None

        return Document(str(path))

    @staticmethod
    def _render_table(table: Any) -> str:
        rows: list[list[str]] = []
        for row in table.rows:
            values = [cell.text.strip() for cell in row.cells]
            if any(values):
                rows.append(values)
        if not rows:
            return ""
        headers = rows[0]
        return _render_markdown_table(headers, rows[1:])


class XLSXTableChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, row_group_size: int = 24) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._row_group_size = max(1, row_group_size)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        path = _metadata_path(metadata, "file_path")
        if path is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        workbook = self._load_workbook(path)
        if workbook is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        for sheet in workbook.worksheets:
            explicit_tables = getattr(sheet, "tables", {}) or {}
            if explicit_tables:
                for table_name, table in explicit_tables.items():
                    rows = self._read_sheet_rows(sheet, table.ref)
                    chunks.extend(self._chunk_rows(source_id, rows, metadata, now, index, sheet.title, table_name))
                    index += len(chunks) - index
            else:
                rows = self._read_sheet_rows(sheet, None)
                chunks.extend(self._chunk_rows(source_id, rows, metadata, now, index, sheet.title, None))
                index += len(chunks) - index

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _load_workbook(path: Path) -> Any | None:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError:
            return None

        return load_workbook(filename=str(path), data_only=True, read_only=True)

    @staticmethod
    def _read_sheet_rows(sheet: Any, table_ref: str | None) -> list[list[str]]:
        from openpyxl.utils.cell import range_boundaries  # type: ignore[import-not-found]

        if table_ref:
            min_col, min_row, max_col, max_row = range_boundaries(table_ref)
            iterator = sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        else:
            iterator = sheet.iter_rows(values_only=True)

        rows: list[list[str]] = []
        for row in iterator:
            values = ["" if value is None else str(value) for value in row]
            if any(cell.strip() for cell in values):
                rows.append(values)
        return rows

    def _chunk_rows(
        self,
        source_id: str,
        rows: list[list[str]],
        metadata: dict[str, object],
        now: datetime,
        start_index: int,
        sheet_title: str,
        table_name: str | None,
    ) -> list[BaseChunk]:
        if not rows:
            return []

        headers = rows[0]
        body_rows = rows[1:] if len(rows) > 1 else []
        chunks: list[BaseChunk] = []
        index = start_index
        for offset in range(0, max(1, len(body_rows)), self._row_group_size):
            group = body_rows[offset : offset + self._row_group_size]
            if not group and body_rows:
                continue
            table_text = _render_markdown_table(headers, group)
            if not table_text.strip():
                continue
            chunk_meta = dict(metadata)
            chunk_meta["sheet"] = sheet_title
            if table_name:
                chunk_meta["table_name"] = table_name
            chunk_meta["row_start"] = offset + 2 if body_rows else 1
            chunk_meta["row_end"] = offset + 1 + len(group) if body_rows else 1
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=table_text,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1
        return chunks


class JSONSubtreeChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_chars: int = 2400) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_chars = max_chars

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        def emit(path: str, value: object) -> None:
            nonlocal index
            rendered = _render_json_node(path, value)
            if not rendered.strip():
                return
            chunk_meta = dict(metadata)
            chunk_meta["json_path"] = path
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=rendered,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        if isinstance(payload, dict):
            for key, value in payload.items():
                emit(str(key), value)
        elif isinstance(payload, list):
            for idx, value in enumerate(payload, start=1):
                emit(f"[{idx}]", value)
        else:
            emit("root", payload)

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)


class XMLSubtreeChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_chars: int = 2400) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_chars = max_chars

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        try:
            root = ET.fromstring(text)
        except ET.ParseError:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        def emit(path: str, node: ET.Element) -> None:
            nonlocal index
            rendered = _render_xml_node(path, node)
            if not rendered.strip():
                return
            chunk_meta = dict(metadata)
            chunk_meta["xml_path"] = path
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=rendered,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        emit(_xml_local_name(root.tag), root)
        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)


class CSVRowGroupChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, row_group_size: int = 24) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._row_group_size = max(1, row_group_size)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        rows = _parse_csv_rows(text)
        if not rows:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        headers = rows[0]
        body_rows = rows[1:] if len(rows) > 1 else []
        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        for offset in range(0, max(1, len(body_rows)), self._row_group_size):
            group = body_rows[offset : offset + self._row_group_size]
            if not group and body_rows:
                continue
            table_text = _render_markdown_table(headers, group)
            if not table_text.strip():
                continue
            chunk_meta = dict(metadata)
            chunk_meta["row_start"] = offset + 2 if body_rows else 1
            chunk_meta["row_end"] = offset + 1 + len(group) if body_rows else 1
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=table_text,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)


class PDFLayoutChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_chars: int = 2500) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_chars = max_chars

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        path = _metadata_path(metadata, "file_path")
        if path is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        document = self._load_document(path)
        if document is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        for page_index, page in enumerate(document, start=1):
            try:
                blocks = page.get_text("blocks", sort=True)
            except Exception:
                blocks = []
            page_lines: list[str] = []
            for block in blocks:
                if len(block) >= 5:
                    block_text = str(block[4]).strip()
                    if block_text:
                        page_lines.append(block_text)
            page_text = "\n".join(page_lines).strip()
            if not page_text:
                continue

            if len(page_text) <= self._max_chars:
                chunk_meta = dict(metadata)
                chunk_meta["page"] = page_index
                chunk_meta["block_count"] = len(page_lines)
                chunks.append(
                    BaseChunk(
                        id=f"{source_id}:{index}",
                        text=page_text,
                        metadata=chunk_meta,
                        source_id=source_id,
                        chunk_index=index,
                        created_at=now,
                    )
                )
                index += 1
                continue

            sub_chunks = self._fallback.chunk(
                source_id=f"{source_id}:page{page_index}",
                text=page_text,
                metadata={**metadata, "page": page_index},
            )
            for sub_chunk in sub_chunks:
                sub_chunk.id = f"{source_id}:{index}"
                sub_chunk.chunk_index = index
                sub_chunk.created_at = now
                sub_chunk.metadata = dict(sub_chunk.metadata)
                sub_chunk.metadata["page"] = page_index
                chunks.append(sub_chunk)
                index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _load_document(path: Path) -> Any | None:
        try:
            import fitz  # type: ignore[import-not-found]
        except ImportError:
            return None

        return fitz.open(str(path))


def _render_json_node(path: str, value: object) -> str:
    lines: list[str] = []

    def walk(current: object, current_path: str) -> None:
        if isinstance(current, dict):
            if not current:
                lines.append(f"{current_path}: {{}}")
                return
            for key, child in current.items():
                next_path = f"{current_path}.{key}" if current_path else str(key)
                walk(child, next_path)
            return
        if isinstance(current, list):
            if not current:
                lines.append(f"{current_path}: []")
                return
            for index, child in enumerate(current, start=1):
                next_path = f"{current_path}[{index}]"
                walk(child, next_path)
            return
        if current_path:
            lines.append(f"{current_path}: {current}")
        else:
            lines.append(str(current))

    walk(value, path)
    return "\n".join(lines)


def _xml_local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    return tag


def _render_xml_node(path: str, node: ET.Element) -> str:
    lines: list[str] = []

    def walk(current: ET.Element, current_path: str) -> None:
        for attr, value in current.attrib.items():
            if value:
                lines.append(f"{current_path}@{attr}: {value}")
        text_value = (current.text or "").strip()
        if text_value:
            lines.append(f"{current_path}: {text_value}")
        for index, child in enumerate(list(current), start=1):
            walk(child, f"{current_path}/{_xml_local_name(child.tag)}[{index}]")

    walk(node, path)
    return "\n".join(lines)


def _parse_csv_rows(text: str) -> list[list[str]]:
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    try:
        dialect = csv.Sniffer().sniff("\n".join(lines[:6]), delimiters=",;\t|")
        rows = list(csv.reader(lines, dialect))
    except csv.Error:
        rows = list(csv.reader(lines))
    normalized = [[str(cell).strip() for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    return normalized


def _heading_level_from_style(style_name: str) -> int | None:
    match = re.search(r"heading\s*(\d+)", style_name, re.IGNORECASE)
    if match:
        return max(1, min(6, int(match.group(1))))
    if style_name.strip().lower() == "title":
        return 1
    return None


def _iter_docx_blocks(document: Any) -> Any:
    try:
        from docx.document import Document as DocxDocument  # type: ignore[import-not-found]
        from docx.table import Table  # type: ignore[import-not-found]
        from docx.text.paragraph import Paragraph  # type: ignore[import-not-found]
    except ImportError:
        return []

    def iter_parent(parent: Any):
        if isinstance(parent, DocxDocument):
            parent_element = parent.element.body
        else:
            parent_element = parent._tc

        for child in parent_element.iterchildren():
            if child.tag.endswith("}p"):
                yield type("DocxBlock", (), {"is_paragraph": True, "paragraph": Paragraph(child, parent)})()
            elif child.tag.endswith("}tbl"):
                yield type("DocxBlock", (), {"is_paragraph": False, "table": Table(child, parent)})()

    return iter_parent(document)


class HtmlDomChunker:
    def __init__(self, fallback: MarkdownAstChunker | None = None, max_chunk_chars: int = 2800) -> None:
        self._fallback = fallback or MarkdownAstChunker()
        self._max_chunk_chars = max(800, max_chunk_chars)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        html_text = self._resolve_html(text=text, metadata=metadata)
        if not html_text.strip():
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        soup = self._load_soup(html_text)
        if soup is None:
            return self._fallback.chunk(source_id=source_id, text=html_text, metadata=metadata)

        root = getattr(soup, "body", None) or soup
        title = self._extract_title(soup)
        if title:
            metadata = dict(metadata)
            metadata.setdefault("title", title)

        chunks: list[BaseChunk] = []
        buffer: list[str] = []
        current_heading_path: list[str] = []
        now = datetime.now(timezone.utc)
        index = 0

        def flush(block_type: str = "paragraph", extra_meta: dict[str, object] | None = None) -> None:
            nonlocal index
            body = "\n\n".join(part.strip() for part in buffer if part.strip()).strip()
            buffer.clear()
            if not body:
                return
            chunk_meta = dict(metadata)
            chunk_meta["document_type"] = DocumentType.HTML.value
            chunk_meta["block_type"] = block_type
            if current_heading_path:
                chunk_meta["heading_path"] = " > ".join(current_heading_path)
            if extra_meta:
                chunk_meta.update(extra_meta)
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=body,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        def emit(text_value: str, block_type: str = "paragraph", extra_meta: dict[str, object] | None = None) -> None:
            nonlocal index
            cleaned = DocumentAwareCleaner().clean(text_value)
            if not cleaned:
                return
            if len(cleaned) > self._max_chunk_chars:
                for piece in self._split_long_text(cleaned, self._max_chunk_chars):
                    emit(piece, block_type=block_type, extra_meta=extra_meta)
                return
            chunk_meta = dict(metadata)
            chunk_meta["document_type"] = DocumentType.HTML.value
            chunk_meta["block_type"] = block_type
            if current_heading_path:
                chunk_meta["heading_path"] = " > ".join(current_heading_path)
            if extra_meta:
                chunk_meta.update(extra_meta)
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=cleaned,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        def visit(node: Any) -> None:
            if not getattr(node, "name", None):
                text_node = str(node).strip()
                if text_node:
                    buffer.append(text_node)
                return

            name = str(node.name).lower()
            if name in {"script", "style", "noscript"}:
                return

            if name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
                flush()
                level = int(name[1])
                heading_text = node.get_text(" ", strip=True)
                if heading_text:
                    while len(current_heading_path) >= level:
                        current_heading_path.pop()
                    current_heading_path.append(heading_text)
                return

            if name == "table":
                flush()
                emit(self._render_table(node), block_type="table")
                return

            if name in {"pre", "code"}:
                flush()
                emit(node.get_text("\n", strip=True), block_type="code")
                return

            if name in {"ul", "ol"}:
                flush()
                emit(self._render_list(node), block_type="list")
                return

            if name in {"p", "blockquote", "li", "figcaption", "caption"}:
                flush()
                emit(node.get_text(" ", strip=True), block_type=name)
                return

            if name in {"article", "section", "div", "main", "header", "footer", "aside", "nav"}:
                for child in node.children:
                    visit(child)
                return

            text_value = node.get_text(" ", strip=True)
            if text_value:
                emit(text_value, block_type=name)

        for child in root.children:
            visit(child)
        flush()
        return chunks or self._fallback.chunk(source_id=source_id, text=html_text, metadata=metadata)

    def _resolve_html(self, text: str, metadata: dict[str, object]) -> str:
        source_html = metadata.get("source_html") or metadata.get("html") or metadata.get("source_body")
        if isinstance(source_html, str) and source_html.strip():
            return source_html

        source_path = _source_path(metadata)
        if source_path and source_path.exists() and source_path.suffix.lower() in {".html", ".htm"}:
            try:
                return source_path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                pass

        return text

    @staticmethod
    def _load_soup(html_text: str) -> Any | None:
        try:
            from bs4 import BeautifulSoup  # type: ignore[import-not-found]
        except ImportError:
            return None

        soup = BeautifulSoup(html_text, "html.parser")
        for tag in soup.find_all(["script", "style", "noscript"]):
            tag.decompose()
        return soup

    @staticmethod
    def _extract_title(soup: Any) -> str:
        title_tag = getattr(soup, "title", None)
        if title_tag is None:
            return ""
        text = title_tag.get_text(" ", strip=True)
        return text.strip()

    @staticmethod
    def _render_list(node: Any) -> str:
        items: list[str] = []
        for li in node.find_all("li", recursive=False):
            item = li.get_text(" ", strip=True)
            if item:
                items.append(f"- {item}")
        return "\n".join(items)

    @staticmethod
    def _render_table(table_node: Any) -> str:
        rows: list[list[str]] = []
        for row in table_node.find_all("tr"):
            cells = row.find_all(["th", "td"])
            values = [cell.get_text(" ", strip=True) for cell in cells]
            if any(value for value in values):
                rows.append(values)
        if not rows:
            return table_node.get_text(" ", strip=True)

        width = max(len(row) for row in rows)
        rows = [row + [""] * (width - len(row)) for row in rows]
        output = [" | ".join(cell.strip() for cell in rows[0])]
        if width > 1:
            output.append(" | ".join("---" for _ in range(width)))
        for row in rows[1:]:
            output.append(" | ".join(cell.strip() for cell in row))
        return "\n".join(output)

    @staticmethod
    def _split_long_text(text: str, max_chars: int) -> list[str]:
        parts: list[str] = []
        cursor = 0
        while cursor < len(text):
            parts.append(text[cursor : cursor + max_chars].strip())
            cursor += max_chars
        return [part for part in parts if part]


class DocxStructureChunker:
    def __init__(self, fallback: MarkdownAstChunker | None = None, max_chunk_chars: int = 3200) -> None:
        self._fallback = fallback or MarkdownAstChunker()
        self._max_chunk_chars = max(800, max_chunk_chars)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        source_path = _source_path(metadata)
        if source_path is None or source_path.suffix.lower() != ".docx" or not source_path.exists():
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        try:
            from docx import Document  # type: ignore[import-not-found]
            from docx.document import Document as DocumentTypeDocx  # type: ignore[import-not-found]
            from docx.table import Table  # type: ignore[import-not-found]
            from docx.text.paragraph import Paragraph  # type: ignore[import-not-found]
        except ImportError:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        document = Document(str(source_path))
        chunks: list[BaseChunk] = []
        buffer: list[str] = []
        heading_path: list[str] = []
        now = datetime.now(timezone.utc)
        index = 0

        def flush(block_type: str = "paragraph", extra_meta: dict[str, object] | None = None) -> None:
            nonlocal index
            body = "\n\n".join(part.strip() for part in buffer if part.strip()).strip()
            buffer.clear()
            if not body:
                return
            chunk_meta = dict(metadata)
            chunk_meta["document_type"] = DocumentType.DOCX.value
            chunk_meta["block_type"] = block_type
            if heading_path:
                chunk_meta["heading_path"] = " > ".join(heading_path)
            if extra_meta:
                chunk_meta.update(extra_meta)
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=body,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        for child in document.element.body.iterchildren():
            if child.tag.endswith("}p"):
                paragraph = Paragraph(child, document)
                text_value = paragraph.text.strip()
                style_name = str(getattr(getattr(paragraph, "style", None), "name", "") or "")
                if style_name.lower().startswith("heading"):
                    flush("heading")
                    level = self._heading_level(style_name)
                    while len(heading_path) >= level:
                        heading_path.pop()
                    if text_value:
                        heading_path.append(text_value)
                    continue
                if text_value:
                    if len(text_value) > self._max_chunk_chars:
                        flush()
                        for piece in self._split_long_text(text_value, self._max_chunk_chars):
                            buffer.append(piece)
                            flush("paragraph")
                    else:
                        buffer.append(text_value)
                else:
                    flush()
            elif child.tag.endswith("}tbl"):
                flush("table")
                table = Table(child, document)
                table_text = self._render_table(table)
                if table_text.strip():
                    chunks.append(
                        BaseChunk(
                            id=f"{source_id}:{index}",
                            text=table_text,
                            metadata={
                                **metadata,
                                "document_type": DocumentType.DOCX.value,
                                "block_type": "table",
                                **({"heading_path": " > ".join(heading_path)} if heading_path else {}),
                            },
                            source_id=source_id,
                            chunk_index=index,
                            created_at=now,
                        )
                    )
                    index += 1

        flush()
        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _heading_level(style_name: str) -> int:
        digits = "".join(ch for ch in style_name if ch.isdigit())
        if digits:
            try:
                return max(1, int(digits))
            except ValueError:
                return 1
        return 1

    @staticmethod
    def _render_table(table: Any) -> str:
        rows: list[list[str]] = []
        for row in table.rows:
            values = [cell.text.strip() for cell in row.cells]
            if any(values):
                rows.append(values)
        if not rows:
            return ""
        width = max(len(row) for row in rows)
        rows = [row + [""] * (width - len(row)) for row in rows]
        output = [" | ".join(rows[0])]
        if width > 1:
            output.append(" | ".join("---" for _ in range(width)))
        for row in rows[1:]:
            output.append(" | ".join(row))
        return "\n".join(output)

    @staticmethod
    def _split_long_text(text: str, max_chars: int) -> list[str]:
        return [text[i : i + max_chars].strip() for i in range(0, len(text), max_chars) if text[i : i + max_chars].strip()]


class XlsxTableChunker:
    def __init__(self, max_rows_per_chunk: int = 24, fallback: FixedWindowChunker | None = None) -> None:
        self._max_rows_per_chunk = max(4, max_rows_per_chunk)
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        source_path = _source_path(metadata)
        if source_path is None or source_path.suffix.lower() != ".xlsx" or not source_path.exists():
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        workbook = load_workbook(filename=str(source_path), data_only=True, read_only=True)
        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            rows = [["" if cell is None else str(cell).strip() for cell in row] for row in rows]
            rows = [row for row in rows if any(cell for cell in row)]
            if not rows:
                continue

            header = self._best_header(rows)
            data_rows = rows[1:] if header is rows[0] else rows
            groups = self._group_rows(data_rows, self._max_rows_per_chunk)
            for group_index, group_rows in enumerate(groups, start=1):
                table_text = self._render_table(header, group_rows)
                if not table_text.strip():
                    continue
                chunk_meta = dict(metadata)
                chunk_meta.update(
                    {
                        "document_type": DocumentType.XLSX.value,
                        "sheet_name": sheet.title,
                        "row_group": group_index,
                        "rows_in_group": len(group_rows),
                    }
                )
                chunks.append(
                    BaseChunk(
                        id=f"{source_id}:{index}",
                        text=table_text,
                        metadata=chunk_meta,
                        source_id=source_id,
                        chunk_index=index,
                        created_at=now,
                    )
                )
                index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _best_header(rows: list[list[str]]) -> list[str]:
        header = rows[0]
        if len(rows) > 1 and any(cell for cell in header):
            return header
        return [f"column_{index}" for index in range(1, max(len(row) for row in rows) + 1)]

    @staticmethod
    def _group_rows(rows: list[list[str]], max_rows: int) -> list[list[list[str]]]:
        if not rows:
            return []
        grouped: list[list[list[str]]] = []
        for index in range(0, len(rows), max_rows):
            grouped.append(rows[index : index + max_rows])
        return grouped

    @staticmethod
    def _render_table(header: list[str], rows: list[list[str]]) -> str:
        width = max(len(header), *(len(row) for row in rows))
        normalized_header = header + [""] * (width - len(header))
        normalized_rows = [row + [""] * (width - len(row)) for row in rows]
        output = [" | ".join(cell or "" for cell in normalized_header)]
        if width > 1:
            output.append(" | ".join("---" for _ in range(width)))
        for row in normalized_rows:
            output.append(" | ".join(cell or "" for cell in row))
        return "\n".join(output)


class JsonSubtreeChunker:
    def __init__(self, fallback: MarkdownAstChunker | None = None, max_chunk_chars: int = 2200) -> None:
        self._fallback = fallback or MarkdownAstChunker()
        self._max_chunk_chars = max(600, max_chunk_chars)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        payload = self._load_json(text=text, metadata=metadata)
        if payload is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        nodes = self._top_level_nodes(payload)
        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        for path, node in nodes:
            lines = self._render_subtree(node, path)
            for part_index, part in enumerate(self._pack_lines(lines, self._max_chunk_chars), start=1):
                chunk_meta = dict(metadata)
                chunk_meta.update({"document_type": DocumentType.JSON.value, "json_path": path, "subtree_part": part_index})
                chunks.append(
                    BaseChunk(
                        id=f"{source_id}:{index}",
                        text=part,
                        metadata=chunk_meta,
                        source_id=source_id,
                        chunk_index=index,
                        created_at=now,
                    )
                )
                index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    def _load_json(self, text: str, metadata: dict[str, object]) -> object | None:
        source_path = _source_path(metadata)
        if source_path and source_path.exists() and source_path.suffix.lower() in {".json", ".jsonl"}:
            try:
                text = source_path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                pass
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return None

    @staticmethod
    def _top_level_nodes(payload: object) -> list[tuple[str, object]]:
        if isinstance(payload, dict):
            return [(str(key), value) for key, value in payload.items()]
        if isinstance(payload, list):
            return [(f"[{index}]", value) for index, value in enumerate(payload, start=1)]
        return [("value", payload)]

    def _render_subtree(self, node: object, path: str) -> list[str]:
        lines: list[str] = []

        def walk(current: object, current_path: str) -> None:
            if isinstance(current, dict):
                if not current:
                    lines.append(f"{current_path}: {{}}")
                    return
                for key, value in current.items():
                    walk(value, f"{current_path}.{key}" if current_path else str(key))
                return
            if isinstance(current, list):
                if not current:
                    lines.append(f"{current_path}: []")
                    return
                for index, value in enumerate(current, start=1):
                    walk(value, f"{current_path}[{index}]")
                return
            lines.append(f"{current_path}: {self._scalar_to_text(current)}")

        walk(node, path)
        return lines

    @staticmethod
    def _scalar_to_text(value: object) -> str:
        if value is None:
            return "null"
        return str(value)

    @staticmethod
    def _pack_lines(lines: list[str], max_chars: int) -> list[str]:
        chunks: list[str] = []
        current: list[str] = []
        size = 0
        for line in lines:
            line_len = len(line) + 1
            if current and size + line_len > max_chars:
                chunks.append("\n".join(current).strip())
                current = []
                size = 0
            current.append(line)
            size += line_len
        if current:
            chunks.append("\n".join(current).strip())
        return [chunk for chunk in chunks if chunk]


class XmlSubtreeChunker:
    def __init__(self, fallback: MarkdownAstChunker | None = None, max_chunk_chars: int = 2200) -> None:
        self._fallback = fallback or MarkdownAstChunker()
        self._max_chunk_chars = max(600, max_chunk_chars)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        root = self._load_xml(text=text, metadata=metadata)
        if root is None:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        for path, node in self._top_level_nodes(root):
            lines = self._render_subtree(node, path)
            for part_index, part in enumerate(self._pack_lines(lines, self._max_chunk_chars), start=1):
                chunk_meta = dict(metadata)
                chunk_meta.update({"document_type": DocumentType.XML.value, "xml_path": path, "subtree_part": part_index})
                chunks.append(
                    BaseChunk(
                        id=f"{source_id}:{index}",
                        text=part,
                        metadata=chunk_meta,
                        source_id=source_id,
                        chunk_index=index,
                        created_at=now,
                    )
                )
                index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    def _load_xml(self, text: str, metadata: dict[str, object]) -> ET.Element | None:
        source_path = _source_path(metadata)
        if source_path and source_path.exists() and source_path.suffix.lower() == ".xml":
            try:
                text = source_path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                pass
        try:
            return ET.fromstring(text)
        except ET.ParseError:
            return None

    @staticmethod
    def _top_level_nodes(root: ET.Element) -> list[tuple[str, ET.Element]]:
        children = list(root)
        if not children:
            return [(XmlSubtreeChunker._local_name(root.tag), root)]
        return [(XmlSubtreeChunker._local_name(child.tag), child) for child in children]

    def _render_subtree(self, node: ET.Element, path: str) -> list[str]:
        lines: list[str] = []

        def walk(current: ET.Element, current_path: str) -> None:
            for attr, value in current.attrib.items():
                if value:
                    lines.append(f"{current_path}@{attr}: {value}")
            text_value = (current.text or "").strip()
            if text_value:
                lines.append(f"{current_path}: {text_value}")
            children = list(current)
            if not children:
                return
            for index, child in enumerate(children, start=1):
                walk(child, f"{current_path}/{self._local_name(child.tag)}[{index}]")

        walk(node, path)
        return lines

    @staticmethod
    def _local_name(tag: str) -> str:
        if "}" in tag:
            return tag.rsplit("}", 1)[1]
        return tag

    @staticmethod
    def _pack_lines(lines: list[str], max_chars: int) -> list[str]:
        chunks: list[str] = []
        current: list[str] = []
        size = 0
        for line in lines:
            line_len = len(line) + 1
            if current and size + line_len > max_chars:
                chunks.append("\n".join(current).strip())
                current = []
                size = 0
            current.append(line)
            size += line_len
        if current:
            chunks.append("\n".join(current).strip())
        return [chunk for chunk in chunks if chunk]


class CsvRowGroupChunker:
    def __init__(self, rows_per_chunk: int = 24, fallback: FixedWindowChunker | None = None) -> None:
        self._rows_per_chunk = max(4, rows_per_chunk)
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        rows = self._load_rows(text=text, metadata=metadata)
        if not rows:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        header = rows[0]
        data_rows = rows[1:] if len(rows) > 1 else []
        if not data_rows:
            data_rows = [header]
            header = [f"column_{index}" for index in range(1, len(header) + 1)]

        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0
        for group_index, start in enumerate(range(0, len(data_rows), self._rows_per_chunk), start=1):
            group_rows = data_rows[start : start + self._rows_per_chunk]
            table_text = self._render_table(header, group_rows)
            if not table_text.strip():
                continue
            chunk_meta = dict(metadata)
            chunk_meta.update(
                {
                    "document_type": DocumentType.CSV.value,
                    "row_group": group_index,
                    "row_start": start + 1,
                    "row_end": start + len(group_rows),
                }
            )
            chunks.append(
                BaseChunk(
                    id=f"{source_id}:{index}",
                    text=table_text,
                    metadata=chunk_meta,
                    source_id=source_id,
                    chunk_index=index,
                    created_at=now,
                )
            )
            index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    def _load_rows(self, text: str, metadata: dict[str, object]) -> list[list[str]]:
        source_path = _source_path(metadata)
        if source_path and source_path.exists() and source_path.suffix.lower() == ".csv":
            try:
                text = source_path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                pass

        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            return []

        try:
            dialect = csv.Sniffer().sniff("\n".join(lines[:6]), delimiters=",;\t|")
            rows = list(csv.reader(lines, dialect))
        except csv.Error:
            rows = list(csv.reader(lines))

        rows = [[cell.strip() for cell in row] for row in rows if any(cell.strip() for cell in row)]
        return rows

    @staticmethod
    def _render_table(header: list[str], rows: list[list[str]]) -> str:
        width = max(len(header), *(len(row) for row in rows))
        normalized_header = header + [""] * (width - len(header))
        normalized_rows = [row + [""] * (width - len(row)) for row in rows]
        output = [" | ".join(cell.strip() for cell in normalized_header)]
        if width > 1:
            output.append(" | ".join("---" for _ in range(width)))
        for row in normalized_rows:
            output.append(" | ".join(cell.strip() for cell in row))
        return "\n".join(output)


class PdfLayoutChunker:
    def __init__(self, fallback: FixedWindowChunker | None = None, max_chunk_chars: int = 3600) -> None:
        self._fallback = fallback or FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._max_chunk_chars = max(1000, max_chunk_chars)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        source_path = _source_path(metadata)
        if source_path is None or source_path.suffix.lower() != ".pdf" or not source_path.exists():
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        try:
            import fitz  # type: ignore[import-not-found]
        except ImportError:
            return self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

        doc = fitz.open(str(source_path))
        chunks: list[BaseChunk] = []
        now = datetime.now(timezone.utc)
        index = 0

        for page_index in range(len(doc)):
            page = doc[page_index]
            blocks = page.get_text("blocks", sort=True) or []
            block_texts: list[str] = []
            for block in blocks:
                if len(block) >= 5:
                    block_text = str(block[4]).strip()
                else:
                    block_text = ""
                if block_text:
                    block_texts.append(DocumentAwareCleaner().clean(block_text))

            if not block_texts:
                fallback_text = DocumentAwareCleaner().clean(page.get_text("text") or "")
                block_texts = [fallback_text] if fallback_text else []

            page_text = "\n\n".join(part for part in block_texts if part.strip()).strip()
            if not page_text:
                continue

            for part_index, part in enumerate(self._split_long_text(page_text, self._max_chunk_chars), start=1):
                chunk_meta = dict(metadata)
                chunk_meta.update(
                    {
                        "document_type": DocumentType.PDF.value,
                        "page_number": page_index + 1,
                        "page_part": part_index,
                        "layout_blocks": len(block_texts),
                    }
                )
                chunks.append(
                    BaseChunk(
                        id=f"{source_id}:{index}",
                        text=part,
                        metadata=chunk_meta,
                        source_id=source_id,
                        chunk_index=index,
                        created_at=now,
                    )
                )
                index += 1

        return chunks or self._fallback.chunk(source_id=source_id, text=text, metadata=metadata)

    @staticmethod
    def _split_long_text(text: str, max_chars: int) -> list[str]:
        return [text[i : i + max_chars].strip() for i in range(0, len(text), max_chars) if text[i : i + max_chars].strip()]


def _source_path(metadata: dict[str, object]) -> Path | None:
    raw = metadata.get("file_path") or metadata.get("source_path") or metadata.get("path")
    if not raw:
        return None
    try:
        return Path(str(raw))
    except Exception:
        return None


class AdaptiveDocumentChunker:
    def __init__(self) -> None:
        from ragflow_orchestrator.chunking.code_python import PythonCodeChunker

        self._code_chunker = PythonCodeChunker()
        self._linear_chunker = FixedWindowChunker(chunk_size=900, chunk_overlap=120)
        self._markdown_chunker = MarkdownAstChunker()
        self._html_chunker = HtmlDomChunker(fallback=self._markdown_chunker)
        self._docx_chunker = DocxStructureChunker(fallback=self._markdown_chunker)
        self._xlsx_chunker = XlsxTableChunker(fallback=self._linear_chunker)
        self._json_chunker = JsonSubtreeChunker(fallback=self._markdown_chunker)
        self._xml_chunker = XmlSubtreeChunker(fallback=self._markdown_chunker)
        self._csv_chunker = CsvRowGroupChunker(fallback=self._linear_chunker)
        self._pdf_chunker = PdfLayoutChunker(fallback=self._linear_chunker)

    def chunk(self, source_id: str, text: str, metadata: dict[str, object] | None = None) -> list[BaseChunk]:
        metadata = dict(metadata or {})
        document_type = self._document_type(metadata=metadata, source_id=source_id, text=text)
        metadata.setdefault("document_type", document_type.value)

        if document_type == DocumentType.CODE:
            return self._code_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.HTML:
            return self._html_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.DOCX:
            return self._docx_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.XLSX:
            return self._xlsx_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.JSON:
            return self._json_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.XML:
            return self._xml_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.CSV:
            return self._csv_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.PDF:
            return self._pdf_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type in {DocumentType.TXT, DocumentType.UNSUPPORTED}:
            return self._linear_chunker.chunk(source_id=source_id, text=text, metadata=metadata)

        if document_type == DocumentType.MARKDOWN:
            normalized = self._normalize_structured_text(text=text, document_type=document_type)
            return self._markdown_chunker.chunk(source_id=source_id, text=normalized, metadata=metadata)

        normalized = self._normalize_structured_text(text=text, document_type=document_type)
        return self._linear_chunker.chunk(source_id=source_id, text=normalized, metadata=metadata)

    @staticmethod
    def _document_type(metadata: dict[str, object], source_id: str, text: str) -> DocumentType:
        explicit = str(metadata.get("document_type") or metadata.get("doctype") or "").strip().lower()
        if explicit and explicit in DocumentType._value2member_map_:
            return DocumentType(explicit)

        source_name = str(metadata.get("file_path") or metadata.get("source_url") or source_id or "")
        content_type = str(metadata.get("content_type") or "").strip().lower()
        detected = detect_document_type(source_name=source_name, content_type=content_type, text=text)
        return detected.document_type

    def _normalize_structured_text(self, text: str, document_type: DocumentType) -> str:
        cleaned = DocumentAwareCleaner().clean(text)
        if document_type == DocumentType.JSON:
            return _normalize_json(cleaned)
        if document_type == DocumentType.XML:
            return _normalize_xml(cleaned)
        if document_type == DocumentType.CSV:
            return _normalize_csv(cleaned)
        return cleaned


def detect_document_type(
    path: Path | None = None,
    *,
    text: str | None = None,
    content_type: str | None = None,
    source_name: str | None = None,
) -> DocumentDetection:
    if content_type:
        normalized_type = _normalize_mime(content_type)
        mapped = _CONTENT_TYPE_MAP.get(normalized_type)
        if mapped:
            return DocumentDetection(document_type=mapped, mime_type=normalized_type, source="content-type")

    if path is not None:
        detected = _detect_from_path(path)
        if detected.document_type != DocumentType.UNSUPPORTED:
            return detected

    if text:
        detected = _detect_from_text(text)
        if detected.document_type != DocumentType.UNSUPPORTED:
            return detected

    if source_name:
        suffix = Path(source_name).suffix.lower()
        mapped = _EXTENSION_MAP.get(suffix)
        if mapped:
            return DocumentDetection(document_type=mapped, source="extension")

    return DocumentDetection(document_type=DocumentType.UNSUPPORTED)


def _detect_from_path(path: Path) -> DocumentDetection:
    fallback = _EXTENSION_MAP.get(path.suffix.lower(), DocumentType.UNSUPPORTED)
    try:
        data = path.read_bytes()[:8192]
    except OSError:
        data = b""

    detected = _detect_from_bytes(data)
    if detected.document_type != DocumentType.UNSUPPORTED:
        return detected

    if fallback != DocumentType.UNSUPPORTED:
        return DocumentDetection(document_type=fallback, source="extension")

    if data:
        return _detect_from_text(data.decode("utf-8", errors="ignore"))
    return DocumentDetection(document_type=DocumentType.UNSUPPORTED)


def _detect_from_bytes(data: bytes) -> DocumentDetection:
    if not data:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    for detector in (_detect_with_filetype, _detect_with_magic):
        detected = detector(data)
        if detected.document_type != DocumentType.UNSUPPORTED:
            return detected

    return DocumentDetection(document_type=DocumentType.UNSUPPORTED)


def _detect_with_filetype(data: bytes) -> DocumentDetection:
    try:
        import filetype  # type: ignore[import-not-found]
    except ImportError:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    kind = filetype.guess(data)
    if kind is None:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    mime = (getattr(kind, "mime", "") or "").lower()
    extension = (getattr(kind, "extension", "") or "").lower()
    mapped = _CONTENT_TYPE_MAP.get(mime)
    if mapped:
        return DocumentDetection(document_type=mapped, mime_type=mime, source="filetype")
    if extension:
        ext_key = f".{extension}"
        if ext_key in _EXTENSION_MAP:
            return DocumentDetection(document_type=_EXTENSION_MAP[ext_key], mime_type=mime, source="filetype")
    return DocumentDetection(document_type=DocumentType.UNSUPPORTED)


def _detect_with_magic(data: bytes) -> DocumentDetection:
    try:
        import magic  # type: ignore[import-not-found]
    except ImportError:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    try:
        mime = str(magic.from_buffer(data, mime=True)).lower().strip()
    except Exception:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    mapped = _CONTENT_TYPE_MAP.get(mime)
    if mapped:
        return DocumentDetection(document_type=mapped, mime_type=mime, source="magic")
    return DocumentDetection(document_type=DocumentType.UNSUPPORTED)


def _detect_from_text(text: str) -> DocumentDetection:
    sample = text.lstrip()
    if not sample:
        return DocumentDetection(document_type=DocumentType.UNSUPPORTED)

    if _looks_like_json(sample):
        return DocumentDetection(document_type=DocumentType.JSON, source="text")
    if _looks_like_html(sample):
        return DocumentDetection(document_type=DocumentType.HTML, source="text")
    if _looks_like_xml(sample):
        return DocumentDetection(document_type=DocumentType.XML, source="text")
    if _looks_like_csv(sample):
        return DocumentDetection(document_type=DocumentType.CSV, source="text")
    if _looks_like_markdown(sample):
        return DocumentDetection(document_type=DocumentType.MARKDOWN, source="text")
    return DocumentDetection(document_type=DocumentType.TXT, source="text")


def _normalize_mime(content_type: str) -> str:
    return content_type.split(";", 1)[0].strip().lower()


def _looks_like_json(text: str) -> bool:
    if not text or text[0] not in "[{":
        return False
    try:
        json.loads(text)
    except json.JSONDecodeError:
        return False
    return True


def _looks_like_xml(text: str) -> bool:
    if not text.startswith("<"):
        return False
    try:
        ET.fromstring(text)
    except ET.ParseError:
        return False
    return True


def _looks_like_html(text: str) -> bool:
    return bool(
        re.search(
            r"<!doctype\s+html|<html\b|</html\b|<head\b|</head\b|<body\b|</body\b|<meta\b|<title\b|<link\b|<script\b|<style\b",
            text,
            re.IGNORECASE,
        )
    )


def _looks_like_csv(text: str) -> bool:
    lines = [line for line in text.splitlines() if line.strip()][:5]
    if len(lines) < 2:
        return False
    sample = "\n".join(lines)
    if not any(delimiter in sample for delimiter in (",", ";", "\t", "|")):
        return False
    try:
        rows = list(csv.reader(lines))
    except csv.Error:
        return False
    return len(rows) >= 2 and max(len(row) for row in rows) > 1


def _looks_like_markdown(text: str) -> bool:
    return any(hint in text for hint in _MARKDOWN_HINTS)


def _normalize_json(text: str) -> str:
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return text

    lines: list[str] = []

    def walk(value: object, prefix: str = "") -> None:
        if isinstance(value, dict):
            if not value and prefix:
                lines.append(f"{prefix}: {{}}")
                return
            for key, child in value.items():
                walk(child, f"{prefix}.{key}" if prefix else str(key))
            return
        if isinstance(value, list):
            if not value and prefix:
                lines.append(f"{prefix}: []")
                return
            for index, child in enumerate(value, start=1):
                walk(child, f"{prefix}[{index}]" if prefix else f"[{index}]")
            return
        lines.append(f"{prefix}: {value}" if prefix else str(value))

    walk(payload)
    return "\n".join(lines) if lines else text


def _normalize_xml(text: str) -> str:
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return text

    lines: list[str] = []

    def local_name(tag: str) -> str:
        if "}" in tag:
            return tag.rsplit("}", 1)[1]
        return tag

    def walk(node: ET.Element, path: str) -> None:
        for attr, value in node.attrib.items():
            if value:
                lines.append(f"{path}@{attr}: {value}")
        node_text = (node.text or "").strip()
        if node_text:
            lines.append(f"{path}: {node_text}")
        for index, child in enumerate(list(node), start=1):
            walk(child, f"{path}/{local_name(child.tag)}[{index}]")

    walk(root, local_name(root.tag))
    return "\n".join(lines) if lines else text


def _normalize_csv(text: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        rows = list(csv.reader(text.splitlines(), dialect))
    except csv.Error:
        rows = list(csv.reader(text.splitlines()))

    if not rows:
        return text

    width = max(len(row) for row in rows)
    rows = [list(row) + [""] * (width - len(row)) for row in rows]

    output = [" | ".join(cell.strip() for cell in rows[0])]
    if width > 1:
        output.append(" | ".join("---" for _ in range(width)))
    for row in rows[1:]:
        output.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(output)
