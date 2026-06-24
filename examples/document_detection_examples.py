"""Examples and integration patterns for universal document detection."""

from pathlib import Path

from ragflow_orchestrator.document_detector import (
    DocumentType,
    UniversalDocumentDetector,
    batch_detect_documents,
    detect_document_type,
    detect_pdf_type,
)

# ============================================================================
# Example 1: Basic detection from file path
# ============================================================================

def example_basic_detection():
    """Detect document type from file path."""
    file_path = Path("regulations.pdf")
    
    # Simple one-liner
    detection = detect_document_type(file_path=file_path)
    
    print(f"Type: {detection.document_type}")
    print(f"Source: {detection.source}")
    print(f"Confidence: {detection.confidence:.1%}")
    
    # For PDF, refine to text vs scan
    if detection.document_type in (DocumentType.PDF_TEXT, DocumentType.PDF_SCAN):
        pdf_detection = detect_pdf_type(file_path)
        print(f"PDF Subtype: {pdf_detection.pdf_subtype}")


# ============================================================================
# Example 2: Detection with file data and text
# ============================================================================

def example_comprehensive_detection():
    """Detect using multiple signals."""
    file_path = Path("document.docx")
    
    # Read file data
    with open(file_path, "rb") as f:
        file_data = f.read()
    
    # Read text (for text formats)
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            text = f.read()
    except UnicodeDecodeError:
        pass
    
    # Create detector
    detector = UniversalDocumentDetector(enable_ocr=False)
    
    # Detect with all signals
    detection = detector.detect(
        file_path=file_path,
        file_data=file_data,
        text=text,
    )
    
    print(f"Detected: {detection.document_type.value}")
    print(f"Confidence: {detection.confidence:.1%}")
    print(f"Method: {detection.source}")


# ============================================================================
# Example 3: PDF classification (text vs scan)
# ============================================================================

def example_pdf_classification():
    """Detect PDF subtype: text layer vs scanned image."""
    pdf_path = Path("regulations.pdf")
    
    # Create detector with OCR enabled
    detector = UniversalDocumentDetector(enable_ocr=True)
    
    # Classify PDF
    detection = detector.detect_pdf_subtype(pdf_path)
    
    if detection.document_type == DocumentType.PDF_TEXT:
        print("✓ PDF has text layer - can extract text directly")
        # Use standard text extraction
    elif detection.document_type == DocumentType.PDF_SCAN:
        print("✓ PDF is scanned - needs OCR processing")
        # Use OCR pipeline
    else:
        print("? Cannot determine PDF type")


# ============================================================================
# Example 4: Custom OCR callable
# ============================================================================

def example_custom_ocr():
    """Use custom OCR detection function."""
    
    def my_ocr_detector(file_path: str) -> bool:
        """Custom function to detect if PDF is scanned.
        
        Returns True if PDF is scanned (needs OCR).
        """
        # Your custom logic here
        # Could call external service, ML model, etc.
        return check_if_scanned_pdf(file_path)
    
    detector = UniversalDocumentDetector(
        enable_ocr=True,
        ocr_callable=my_ocr_detector,
    )
    
    detection = detector.detect_pdf_subtype("regulations.pdf")
    print(f"PDF type: {detection.pdf_subtype}")


def check_if_scanned_pdf(file_path: str) -> bool:
    """Placeholder for custom OCR logic."""
    # Your implementation
    return False


# ============================================================================
# Example 5: Batch detection for multiple files
# ============================================================================

def example_batch_detection():
    """Detect types for folder of documents."""
    documents_dir = Path("documents")
    
    # Get all files
    file_paths = list(documents_dir.glob("*"))
    
    # Batch detect
    results = batch_detect_documents(file_paths)
    
    # Group by type
    by_type = {}
    for file_path, detection in results.items():
        doc_type = detection.document_type.value
        if doc_type not in by_type:
            by_type[doc_type] = []
        by_type[doc_type].append(file_path)
    
    # Print summary
    for doc_type, files in sorted(by_type.items()):
        print(f"\n{doc_type.upper()} ({len(files)} files):")
        for f in files:
            print(f"  - {Path(f).name}")


# ============================================================================
# Example 6: Integration with ingestion pipeline
# ============================================================================

def example_pipeline_integration():
    """Use detector in document ingestion pipeline."""
    from ragflow_orchestrator.versioned_pipeline import VersionedDocumentPipeline
    
    # Create detector
    detector = UniversalDocumentDetector(enable_ocr=True)
    
    # In your pipeline:
    def process_document(file_path: Path, pipeline: VersionedDocumentPipeline):
        """Process document with automatic type detection."""
        
        # 1. Detect type
        with open(file_path, "rb") as f:
            file_data = f.read()
        
        detection = detector.detect(file_path=file_path, file_data=file_data)
        
        print(f"Processing {file_path.name} as {detection.document_type.value}")
        
        # 2. Read content
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
        except UnicodeDecodeError:
            # Binary file - extract text using appropriate tool
            text = extract_text_from_binary(file_path, detection.document_type)
        
        # 3. Ingest with detected type
        result = pipeline.ingest_document(
            file_path=file_path,
            text=text,
            document_type=map_detected_type(detection.document_type),
            auto_tags=[detection.document_type.value],
        )
        
        return result


# ============================================================================
# Example 7: Confidence-based routing
# ============================================================================

def example_confidence_routing():
    """Route processing based on detection confidence."""
    file_path = Path("unknown_document.bin")
    
    detector = UniversalDocumentDetector(enable_ocr=False)
    detection = detector.detect(file_path=file_path)
    
    if detection.confidence > 0.9:
        # High confidence - use detected type directly
        process_as_type(detection.document_type)
    elif detection.confidence > 0.7:
        # Medium confidence - may need verification
        process_with_verification(detection.document_type)
    else:
        # Low confidence - ask user or use manual inspection
        ask_user_for_document_type(file_path)


def process_as_type(doc_type: DocumentType):
    print(f"Processing as {doc_type.value}")


def process_with_verification(doc_type: DocumentType):
    print(f"Processing as {doc_type.value} (verification recommended)")


def ask_user_for_document_type(file_path: Path):
    print(f"Cannot auto-detect {file_path.name}. Please specify type:")


# ============================================================================
# Example 8: Detailed detection information
# ============================================================================

def example_detailed_info():
    """Get full detection information."""
    file_path = Path("regulations.pdf")
    
    detector = UniversalDocumentDetector(enable_ocr=True)
    
    # Read file
    with open(file_path, "rb") as f:
        file_data = f.read()
    
    # Detect
    detection = detector.detect(file_path=file_path, file_data=file_data)
    
    # Print detailed info
    print(f"Document Type: {detection.document_type.value}")
    print(f"MIME Type: {detection.mime_type or 'N/A'}")
    print(f"Detection Source: {detection.source}")
    print("  - extension: fast check based on file extension")
    print("  - magic_bytes: binary signature analysis")
    print("  - mime: MIME type from python-magic")
    print("  - content: text pattern matching")
    print(f"Confidence: {detection.confidence:.1%}")
    print(f"PDF Subtype: {detection.pdf_subtype or 'N/A'}")
    
    # Interpretation
    if detection.confidence > 0.9:
        certainty = "Very high certainty"
    elif detection.confidence > 0.7:
        certainty = "Good confidence"
    else:
        certainty = "Low confidence - verification recommended"
    print(f"\nCertainty: {certainty}")


# ============================================================================
# Helper functions for integration
# ============================================================================

def map_detected_type(detected_type: DocumentType):
    """Map UniversalDocumentDetector type to VersionedDocumentPipeline type."""
    # Import here to avoid circular dependency
    from ragflow_orchestrator.document_pipeline import DocumentType as PipelineDocumentType
    
    mapping = {
        DocumentType.DOCX: PipelineDocumentType.DOCX,
        DocumentType.XLSX: PipelineDocumentType.XLSX,
        DocumentType.PDF_TEXT: PipelineDocumentType.PDF,
        DocumentType.PDF_SCAN: PipelineDocumentType.PDF,
        DocumentType.MARKDOWN: PipelineDocumentType.MARKDOWN,
        DocumentType.JSON: PipelineDocumentType.JSON,
        DocumentType.CSV: PipelineDocumentType.CSV,
        DocumentType.XML: PipelineDocumentType.XML,
        DocumentType.HTML: PipelineDocumentType.HTML,
        DocumentType.PYTHON: PipelineDocumentType.CODE,
        DocumentType.JAVASCRIPT: PipelineDocumentType.CODE,
        DocumentType.TYPESCRIPT: PipelineDocumentType.CODE,
        DocumentType.CODE: PipelineDocumentType.CODE,
    }
    return mapping.get(detected_type, PipelineDocumentType.TXT)


def extract_text_from_binary(file_path: Path, doc_type: DocumentType) -> str:
    """Extract text from binary formats."""
    if doc_type == DocumentType.DOCX:
        return extract_docx_text(file_path)
    elif doc_type == DocumentType.XLSX:
        return extract_xlsx_text(file_path)
    elif doc_type in (DocumentType.PDF_TEXT, DocumentType.PDF_SCAN):
        return extract_pdf_text(file_path, doc_type)
    else:
        return ""


def extract_docx_text(file_path: Path) -> str:
    """Extract text from DOCX."""
    try:
        from docx import Document
        doc = Document(file_path)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)
    except Exception:
        return ""


def extract_xlsx_text(file_path: Path) -> str:
    """Extract text from XLSX."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        text = []
        for sheet in wb.sheetnames:
            text.append(f"\n=== Sheet: {sheet} ===\n")
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text.append(" | ".join(str(v or "") for v in row))
        return "\n".join(text)
    except Exception:
        return ""


def extract_pdf_text(file_path: Path, doc_type: DocumentType) -> str:
    """Extract text from PDF."""
    if doc_type == DocumentType.PDF_TEXT:
        return extract_pdf_text_layer(file_path)
    else:
        return extract_pdf_with_ocr(file_path)


def extract_pdf_text_layer(file_path: Path) -> str:
    """Extract text from PDF text layer."""
    try:
        import fitz
        doc = fitz.open(str(file_path))
        text = []
        for page in doc:
            text.append(page.get_text())
        doc.close()
        return "\n".join(text)
    except Exception:
        return ""


def extract_pdf_with_ocr(file_path: Path) -> str:
    """Extract text from PDF using OCR."""
    try:
        from paddleocr import PaddleOCR
        ocr = PaddleOCR(lang="ru")
        result = ocr.ocr(str(file_path), cls=True)
        text = []
        for line in result:
            for item in line:
                text.append(item[1][0])
        return "\n".join(text)
    except Exception:
        return ""


# ============================================================================
# Main
# ============================================================================

if __name__ == "__main__":
    print("Document Detection Examples\n")
    
    print("=" * 60)
    print("Example 1: Basic Detection")
    print("=" * 60)
    try:
        example_basic_detection()
    except FileNotFoundError:
        print("(Skipped - example files not found)")
    
    print("\n" + "=" * 60)
    print("Example 8: Detailed Information")
    print("=" * 60)
    
    # Create a test file
    test_file = Path("test_detection.json")
    test_file.write_text('{"key": "value"}')
    
    example_detailed_info()
    test_file.unlink()
